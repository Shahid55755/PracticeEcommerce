import openpyxl


class HomePageData:
    @staticmethod
    def gettestdata():
        data_list = []
        book = openpyxl.load_workbook("C:\\Users\\HP\\Desktop\\Datadrivertest\\signupdata.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1, column=2)
        for i in range(2, sheet.max_row + 1):
            data_dict = {}
            for j in range(2, sheet.max_column + 1):
                data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data_list.append(data_dict)
        return data_list

