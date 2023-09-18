import openpyxl


class HomePageData1:
    @staticmethod
    def gettestdata(TestCase3):
        data_list = []
        book = openpyxl.load_workbook("C:\\Users\\HP\\Desktop\\Datadrivertest\\signupdata.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1, column=2)
        for i in range(2, sheet.max_row + 1):
            data_dict = {}  # create dictionary for each row
            if sheet.cell(row=i, column=1).value == TestCase3:  # search entry in column 1, iterate through all rows
                for j in range(2, sheet.max_column + 1):  # iterate from second column to maximum column for data
                    data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data_list.append(data_dict)  # list append the dictionary and return the list
        return data_list

